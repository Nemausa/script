import re

from numpy.lib.utils import source
from stock import Stock
import openpyxl
from openpyxl.styles import PatternFill  # 导入填充模块

stock = Stock()
# 863,002415.SZ,002415,海康威视,浙江,电器仪表,20100528
# 2607,600309.SH,600309,万华化学,山东,化工原料,20010105
# 801,002352.SZ,002352,顺丰控股,深圳,仓储物流,20100205
ts_code = '002415.SZ'
file_name = ts_code + '.xlsx'

df = stock.pro.daily(ts_code=ts_code, start_date='20200101', end_date='20210210')
df.to_excel(file_name)
wb = openpyxl.load_workbook(file_name)
source = wb['Sheet1']

Color = ['8B0000', '3CB371','FFFFE0', '4169E1']  # 深红色 春天的绿色 浅黄色 皇家蓝

column = 7
rise = False

value = source.cell(row=source.max_row, column=column).value
max_value = value
min_value = value
for row in range(source.max_row,1,-1):
    value_pre = 0
    value_next =0

    if row>2:
        value_pre = source.cell(row=row-1, column=column).value
    if row<source.max_row:
        value_next = source.cell(row=row+1, column=column).value

    value = source.cell(row=row, column=column).value

    if value>1.06*min_value and rise==False:
        max_value = min_value
        rise = True
    if value<0.94*max_value and rise==True:
        min_value = max_value
        rise = False

    if value>min_value*1.06 and value>max_value and rise==True:
        max_value = value
        fille = PatternFill('solid', fgColor=Color[0]) 
        source.cell(row=row,column=column).fill = fille

    if value<max_value*0.94 and value<min_value and rise==False:
        min_value = value
        fille = PatternFill('solid', fgColor=Color[1]) 
        source.cell(row=row,column=column).fill = fille
        
wb.save(file_name)

print('end')
