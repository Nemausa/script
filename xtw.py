import re
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



class Nemausa:
    
    def __init__(self) -> None:
        self.first_row=0
        self.data = list()
        self.wb = Workbook()
        self.dest_filename = 'empty_book.xlsx'
        self.ws1 = self.wb.create_sheet(title="Data")

    def read(self, input):
        # self.data.clear()
        with open(input, 'r', encoding='UTF-8') as f:
            lines = f.readlines()
            self.get(lines)

    def get(self, lines):
        pat = re.compile(r'.*>(.*)<.*')
        for line in lines:
            m = pat.match(line)
            if m is not None:
                str = m.group(1)
                # str = re.sub("[A-Za-z0-9\!\%\[\]\,\。]", "", str)
                if str is not '':
                    self.data.append(str)

    def write(self):
        l2 = list(set(self.data))
        l2.sort(key=self.data.index)
        for row in range(l2.__len__()):
            self.first_row += 1
            self.ws1.cell(column=1, row=self.first_row, value=str(l2[row]))
        self.wb.save(filename=self.dest_filename)



count = 0
nemausa = Nemausa()

g = os.walk('xtw')
for path, dir_list, file_list in g:
    for file_name in file_list:
        file = os.path.join(path, file_name)
        nemausa.read(file)
        count +=1
        print(count)


nemausa.write()
os.system('pause')
