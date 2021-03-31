import os
import re
import time
import openpyxl
#from numba import jit



class excel_to_xml():
    ''''''
    def __init__(self) -> None:
        super().__init__()

    def read_excel(self, column):
        filename = 'E3.xlsx'
        wb =  openpyxl.load_workbook(filename = filename)
        self.source = wb['EVR']
        print(self.source.max_row)
        # read value by row
        column = 4
        keys = list()
        values = list()
        for row in range(2,870):
            key = self.source.cell(row=row, column=2).value
            value = self.source.cell(row=row, column=column).value
            keys.append(key)
            values.append(value)
        self.res = {keys[i]: values[i] for i in range(len(keys))} 
        self.keys = keys
        self.values = values

        # print(self.res)


    def mkdir(self, path):
        filename = os.path.basename(path)
        dir_path = path.replace(filename, '')
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)

    def write_xml(self,input):
        # read value by row
        keys = list()
        values = list()

        outdir = 'nemausa/'+ '/'
        outdir= outdir.replace('\n', '')
        self.mkdir(outdir + input)
            
        with open(input, 'r', encoding='UTF-8') as f:
            lines = f.readlines()
            with open(outdir+input, 'w', encoding='UTF-8') as w:
                replace_lines = self.replace(lines)
                w.writelines(replace_lines)

    def replace(self, lines):
        pat = re.compile(r'(?<=>).*(?=<)')
        replace_lines = []
        for line in lines:
            m = pat.findall(line)
            if len(m):
                for i in range(len(self.keys)):
                    if self.keys[i] == m[0]:
                        line = line.replace(self.keys[i], self.values[i])

            else:
                pass
            replace_lines.append(line)
        return replace_lines


begin = time.time()
nemausa = excel_to_xml()
nemausa.read_excel(4)
nemausa.write_xml('strings.xml')
update = time.time() - begin
print('time:' + str(update))
print('finish')
