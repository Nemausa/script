import os
import csv
import time
from xml.etree.ElementTree import fromstring
import openpyxl
import webbrowser
import numpy as np
import tushare as ts
from math import floor
from concurrent.futures import ThreadPoolExecutor, as_completed



def partition(ls, size):
    """
    Returns a new list with elements
    of which is a list of certain size.

        >>> partition([1, 2, 3, 4], 3)
        [[1, 2, 3], [4]]
    """
    return [ls[i:i+size] for i in range(0, len(ls), size)]


class Stock:
    def __init__(self) -> None:
        super().__init__()

        # array = [10*n for n in range(1,2)]
        array = []
        array.append(100)
        self.array = array
        ts.set_token('e0971dcf036c61bd699ac86762b33d9e6d44025979ac91f6dfc58b31')
        self.pro = ts.pro_api()
        # data = self.pro.stock_basic(exchange='', list_status='L', fields='ts_code,symbol,name,area,industry,list_date')
        # data.to_excel('stock.xlsx')
        self.pool = ThreadPoolExecutor(128)
        self.price = list()
        self.result = list()
        self.count = 0



    def read_csv(self):
        csv_file = open('stock.csv', 'r', encoding='UTF-8')
        reader = csv.reader(csv_file)

        data = []
        dic_name = dict()
        for item in reader:
            if reader.line_num == 1:
                continue
            data.append(item[1])
            dic_name[item[1]] = item[3]

        csv_file.close()
        self.data = data
        self.dic_name = dic_name
        

    def find_nearest(self, value):
        array = np.asarray(self.array)
        idx = (np.abs(array - value)).argmin()
        return array[idx]


    def daily(self,ts_code):
        date = '20210202'
        df = self.pro.daily(ts_code=ts_code, start_date=date, end_date=date)
        # print(df)
        # print(df.to_dict())

        # print(df)
        df.to_csv('megrez.csv')
        return df



    def all(self):
        n = 50
        i=0
        res = partition(self.data, floor(len(self.data)/n))

        all_task = [self.pool.submit(self.calcualte, (item)) for item in res]
        for future in as_completed(all_task):
            pass

        

    def calcualte(self, item): 
        df = self.daily(','.join(item)) 
        for value, ts_code, change in zip(df.high, df.ts_code, df.change):
            try:
                # print(ts_code, value)
                near_value = self.find_nearest(value)
                if value>=near_value*0.85 and value<=near_value:
                    # if change > 0:
                    self.result.append(ts_code)
            except:
                pass

            
        # print(self.result)


    def add_name(self):
        openpyxl.load_workbook('stock.xlsx')
        stock_wb =  openpyxl.load_workbook(filename = 'stock.xlsx')
        stock_source = stock_wb['Sheet1']
        dic = dict()
        for row in range(1,stock_source.max_row+1):
            key = stock_source.cell(row=row, column=2).value
            value = stock_source.cell(row=row, column=4).value

            dic[key] = value

        for i in range(self.count+1):
            filename = str(i)+'.xlsx'
            daily_wb = openpyxl.load_workbook(filename=filename)
            daily_source = daily_wb['Sheet1']
            for row in range(2, daily_source.max_row+1):
                key = daily_source.cell(row=row, column=2).value
                daily_source.cell(row=row, column=13, value=dic[key])

            daily_wb.save(filename=filename)   



    def open_url(self):
        for i in range(self.count+1):
            filename = str(i)+'.xlsx'
            daily_wb = openpyxl.load_workbook(filename=filename)
            daily_source = daily_wb['Sheet1']
            for row in range(2, daily_source.max_row+1):
                ts_code = daily_source.cell(row=row, column=2).value
                li_split = ts_code.split('.')
                scode = li_split[-1]+li_split[0]
                url = 'https://xueqiu.com/S/'+scode
                new = 0
                if row%10==0:
                    new = 1
                
                webbrowser.open(url, new)


    def single(self,ts_code):
        begin_date = '20160101'
        end_date = '20210202'
        df=self.pro.daily(ts_code=ts_code, start_date=begin_date, end_date=end_date)       
        df.to_excel('nemausa/'+ts_code+'.xlsx')

    def analysis(self):
        openpyxl.load_workbook('0.xlsx')
        stock_wb =  openpyxl.load_workbook(filename = '0.xlsx')
        stock_result = stock_wb['Sheet1']
      
        ts_code = [stock_result.cell(row=row, column=2).value   for row in range(2, stock_result.max_row+1)]
            

        all_task = [self.pool.submit(self.single, (word)) for word in ts_code]
        for future in as_completed(all_task):
            pass

def max_price():
    g = os.walk('nemausa')
    for path,dir_list,file_list in g:  
        for file_name in file_list:  
            filename=path+'/'+file_name
            wb = openpyxl.load_workbook(filename=filename)
            source = wb['Sheet1']
            L = [source.cell(row=row, column=5).value for row in range(2, source.max_row+1)]
            if max(L) > 100:
                os.remove(filename)

def run():

    time_start = time.time()

    stock = Stock()
    stock.read_csv()
    stock.all()
    print('len: ' + str(len(stock.result)))
    num = floor(len(stock.result)/100)+1
    res = partition(stock.result, floor(len(stock.result)/num))
    for i, item in zip(range(num),res):
        df = stock.daily(','.join(item))
        df.to_excel(str(i)+'.xlsx')
        stock.count = i

    stock.count = 0
    stock.add_name()
    stock.analysis()
    # stock.open_url()

    max_price()
    
    time_end = time.time()
    print('time cost', time_end-time_start, 's')

    print('finish')
    # os.system('pause')


# run()


from datetime import datetime
stock = Stock()
# 863,002415.SZ,002415,海康威视,浙江,电器仪表,20100528
# 2607,600309.SH,600309,万华化学,山东,化工原料,20010105
# 801,002352.SZ,002352,顺丰控股,深圳,仓储物流,20100205
ts_code = '002415.SZ'
df = stock.pro.daily(ts_code=ts_code, start_date='20190101', end_date='20210204')
df.to_excel(ts_code + '.xlsx')
# analysis Monday and Friday about price and chage
wb = openpyxl.load_workbook(ts_code+'.xlsx')
friday_up =0
friday_down = 0
monday_up = 0
monday_down = 0
bill = []
all = []
result = wb['Sheet1']
check = 0
weeks=['星期一','星期二','星期三','星期四','星期五']
l = [x*0 for x in range(10)]
for row in range(2, result.max_row+1):
    date = result.cell(row=row, column=3).value
    change = result.cell(row=row, column=10).value

    week=datetime.strptime(date, "%Y%m%d").weekday()
    result.cell(row=row, column=13, value=weeks[week])
    all.append(change)
    if change<0:
        l[week*2] += 1
    else:
        l[week*2+1] += 1
    
    if week ==1 :
        bill.append(change)

    if week ==0 or week==4:
        bill.append(change)  


base = 100
price = 100
reversd_data = list(reversed(bill))
reversd_all = list(reversed(all))
for a in reversd_data:
    base *=(1+0.01*a)


for a in reversd_all:
    price *=(1+0.01*a)
wb.save(ts_code+'.xlsx')

    
print('stock operation on Friday and Monday:' ,base)
print('stock operation on:' ,price)
for week in range(5):
    print(weeks[week],'up: ', str(l[week*2+1]))
    print(weeks[week],'down: ', str(l[week*2]))






