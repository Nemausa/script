import os
import re
import time
import openpyxl
from openpyxl.utils import cell, get_column_letter
from concurrent.futures import ThreadPoolExecutor,  as_completed
from google_trans_new import google_translator
from collections import defaultdict





class Nemausa:
    
    def __init__(self) -> None:
        self.first_row=0
        self.data = list()
        tree = lambda: defaultdict(tree)
        self.root = tree()
        self.lan = ['fr', 'en', 'ru', 'es', 'tr', 'ar', 'id', 'de', 'th', 'fa', 'it', 'ko', 'nb', 'iw', 'sv', 'bl']
        for key1, key2 in zip(range(3, 19), self.lan):
            self.root[key1][key2] = dict()
        # self.root[3]['fr'] = dict()
        # self.root[4]['en'] = dict()
        # self.root[5]['ru'] = dict()
        # self.root[6]['es'] = dict()
        # self.root[7]['tr'] = dict()
        # self.root[8]['ar'] = dict()
        # self.root[9]['id'] = dict()
        # self.root[10]['de'] = dict()
        # self.root[11]['th'] = dict()
        # self.root[12]['fa'] = dict()
        # self.root[13]['it'] = dict()
        # self.root[14]['kr'] = dict()
        # self.root[15]['nb'] = dict()
        # self.root[16]['he'] = dict()
        # self.root[17]['sv'] = dict()
        # self.root[18]['bl'] = dict()
        self.key1 = None
        self.key2 = None

        self.pool = ThreadPoolExecutor(1000)
        self.translator = google_translator()
        self.wb =  openpyxl.load_workbook(filename = 'E3.xlsx')
        self.source = self.wb['EVR']
        self.filename = 'E3.xlsx'

    def read(self, input):
        self.data.clear()
        self.source = self.wb['EVR']
        print(self.source.max_row)
        # read value by row
        for row in range(2,860):
            key = self.source.cell(row=row, column=2).value
            self.data.append(key)



        # translate
        self.translate_all()
        self.write();

    def write(self):
        for row in range(2,860):
            for key1, key2 in zip(range(3, 19), self.lan): 
                key = str(self.source.cell(row=row, column=2).value)
                try:
                    self.source.cell(column=key1, row=row, value=self.root[key1][key2][key])
                except:
                    pass
        
        self.wb.save(filename=self.filename)

    def translate(self, word):
        translate_text = self.translator.translate(word, lang_tgt=self.key2)
        self.root[self.key1][self.key2][word] = translate_text
        return translate_text

    def translate_all(self):
        
        for key1, key2 in zip(range(3, 19), self.lan): 
            self.key1 = key1
            self.key2 = key2
            all_task = [self.pool.submit(self.translate, (word)) for word in self.data]
            for future in as_completed(all_task):
                pass  
        
          
        
        

def run():
    time_start = time.time()

    
    nemausa = Nemausa()
    nemausa.read(nemausa.filename)
    
    time_end = time.time()
    print('time cost', time_end-time_start, 's')

    print('finish')
    os.system('pause')


run()
